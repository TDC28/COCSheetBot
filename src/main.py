from Client import Client
from Clan import Clan
from dotenv import load_dotenv
from discord import File
from discord.ext import commands
from openpyxl import load_workbook
import os, discord, logging, json


# TODO: Rewrite docstrings for all functions and try to setup autosuggestions from Discord
def run(client, token, filename):
    intents = discord.Intents.all()

    bot = commands.Bot(command_prefix="!", intents=intents)

    @bot.event
    async def on_ready():
        logging.info("User: %s", bot.user)

    @bot.command(name="load_main")
    async def load_main(ctx, clantag):
        """
        load_main loads the CWL data from cocAPI. Compiles the CWL spreadsheet and posts it to the current chat.
        """
        clan = Clan(clantag, client)
        clan.get_data()
        clan.update_sheet(filename)

        await ctx.send("**CWL 2024 Season**")
        await ctx.send(
            "Updated: "
            + clan.wars[0]["startTime"][4:6]
            + "-"
            + clan.wars[0]["startTime"][0:4]
        )
        await ctx.send(file=File(filename))

    @bot.command(name="link_player")
    async def link_player(ctx, account_name, coc_tag):
        """
        link_player links account_name to a valid Clash of Clans player tag.

        :param account_name: Account name to be link to Clash of Clans tag
        :param coc_tag: Clash of Clans player tag
        """
        if coc_tag[0] != "#":
            await ctx.send("Invalid player tag")
            return

        with open("playerList.json", "r") as playerDict:
            linked_players = json.load(playerDict)

        if account_name in linked_players["Players"]:
            await ctx.send("Discord user already linked, updating COC player tag")

        with open("playerList.json", "w") as playerDict:
            linked_players["Players"][account_name] = coc_tag
            json.dump(linked_players, playerDict)

        await ctx.send("Linked **" + account_name + f"** to **{coc_tag}** successfully")

    @bot.command(name="unlink_player")
    async def unlink_player(ctx, account_name):
        """
        unlink_player unlinks account_name from a Clash of Clans player tag.

        :param account_name: Account to be unlinked
        """
        with open("playerList.json", "r") as playerDict:
            linked_players = json.load(playerDict)

        if account_name not in linked_players["Players"]:
            await ctx.send("Discord user is not linked")
            return

        with open("playerList.json", "w") as playerDict:
            linked_players["Players"].pop(account_name)
            json.dump(linked_players, playerDict)

        await ctx.send(f"Unlinked **{account_name}** successfully")

    @bot.command(name="check_link")
    async def check_link(ctx, account_name):
        """
        check_link checks if account_name is linked to a Clash of Clans player tag. If player is linked, outputs linked Clash of Clans player tag.

        :param account_name: Account to be checked
        """
        with open("playerList.json", "r") as playerDict:
            linked_players = json.load(playerDict)

        await ctx.send("linked players")
        await ctx.send(linked_players)

        if account_name not in linked_players["Players"]:
            await ctx.send(account_name + " is not linked")
            return

        await ctx.send(
            account_name
            + f" is currently linked to {linked_players['Players'][account_name]}"
        )

    @bot.command(name="input_player_score")
    async def input_player_score(ctx, account_name, stars, destruction):
        """
        input_player_score add to the spreadsheet the total stars and destruction% earned throughout the CWL to the given account_name.

        :param account_name: Account to add start and destruction stats
        :param stars: Total stars earned by given account in the CWL
        :param destruction: Total destruction% earned by given account in the CWL
        """
        with open("playerList.json", "r") as playerDict:
            linked_players = json.load(playerDict)

        if account_name not in linked_players["Players"]:
            await ctx.send(account_name + " is not linked")
            return

        wb = load_workbook(filename)
        ws = wb["RegisteredPlayers"]
        search = linked_players["Players"][account_name]
        loc = None

        for row in range(200):
            if ws["F" + str(row + 3)].value == search:
                loc = str(row + 3)
                break

        if loc is None:
            await ctx.send(account_name + " is not registered in the sheet")
            return

        new_stars = int(ws["H" + loc].value) + int(stars)
        new_dest = int(ws["I" + loc].value) + int(destruction)

        ws["H" + loc] = new_stars
        ws["I" + loc] = new_dest

        wb.save(filename)

        await ctx.send(account_name + " stats were successfully loaded")

    @bot.command(name="list_players")
    async def list_players(ctx):
        """
        list_players lists all linked players
        """
        with open("playerList.json", "r") as playerDict:
            linked_players = json.load(playerDict)

        links = ""
        for player in linked_players["Players"]:
            links = links + f"{player} ---> {linked_players['Players'][player]}\n"

        await ctx.send(links)

    if TOKEN is not None:
        bot.run(TOKEN)

    else:
        print("Invalid token")
        return


if __name__ == "__main__":
    load_dotenv()
    API_KEY = str(os.environ.get("API_KEY"))
    TOKEN = str(os.environ.get("TOKEN"))
    client = Client(API_KEY)
    filename = "CWL2024.xlsx" # TODO: Rename this to the filename you chose for your file

    run(client, TOKEN, filename)

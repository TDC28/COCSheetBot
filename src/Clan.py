from openpyxl import load_workbook
from Client import Client


class Clan:
    def __init__(self, clantag: str, client: Client) -> None:
        self.client = client
        self.clantag = clantag

        self.des = []
        self.stars = []
        self.member_names = []
        self.member_tags = []
        self.wars = []

        self.league_group = self.client.league_group(self.clantag)

    def clan_in_war(self, wartag: str) -> bool:
        """
        clan_in_war(wartag) produces True if the clan is participating in
            the war with a war tag wartag, else False.
        clan_in_war: Clan str -> bool
        """
        war = self.client.war(wartag)
        clans = (war["clan"]["tag"], war["opponent"]["tag"])

        return not (
            war == {"state": "notInWar", "teamSize": 0} or self.clantag not in clans
        )

    def get_data(self) -> None:
        """
        get_data() populates the lists in the Clan using data from Clash of
            Clans API.
        get_data: Clan -> None
        """
        for group in self.league_group["rounds"]:
            for wartag in group["warTags"]:
                if self.clan_in_war(wartag):
                    self.wars.append(self.client.war(wartag))

        for war in self.wars:
            for i in range(self.wars[0]["teamSize"]):
                if war["clan"]["tag"] == self.clantag:
                    self.member_names.append(war["clan"]["members"][i]["name"])
                    self.member_tags.append(war["clan"]["members"][i]["tag"])

                else:
                    self.member_names.append(war["opponent"]["members"][i]["name"])
                    self.member_tags.append(war["opponent"]["members"][i]["tag"])

        for war in self.wars:
            team_size = war["teamSize"]
            for i in range(team_size):
                if war["clan"]["tag"] == self.clantag:
                    member = war["clan"]["members"][i]

                else:
                    member = war["opponent"]["members"][i]

                if "attacks" in member:
                    attack = member["attacks"][0]
                    self.stars.append(attack["stars"])
                    self.des.append(attack["destructionPercentage"])

                else:
                    self.stars.append("Missed Hit")
                    self.des.append(0)

        return None

    def update_sheet(self, filename: str) -> None:
        """
        update_sheet(filename) edits the Excel spreadsheet.
        update_sheet: Clan str -> None
        """
        wb = load_workbook(filename)
        ws = wb[self.wars[0]["startTime"][4:6]]

        for row in range(len(self.member_tags)):
            ws["A" + str(row + 3)] = self.member_tags[row]
            ws["B" + str(row + 3)] = self.stars[row]
            ws["C" + str(row + 3)] = self.des[row]
            ws["D" + str(row + 3)] = self.member_names[row]

        ws["E3"] = self.wars[0]["teamSize"]

        for row in range(len(list(set(self.member_tags)))):
            ws["G" + str(row + 3)] = list(set(self.member_tags))[row]
            ws["H" + str(row + 3)] = (
                "=_xlfn.XLOOKUP(G"
                + str(row + 3)
                + ",A2:A"
                + str(len(self.member_tags) + 1)
                + ",D2:D"
                + str(len(self.member_tags) + 1)
                + ")"
            )

        for row in range(len(set(self.member_tags)) + 3, 101):
            ws.row_dimensions[row].hidden = True

        ws["S101"] = "=SUM(S3:S" + str(len(list(set(self.member_tags))) + 2) + ")"
        ws["T101"] = "=SUM(T3:T" + str(len(list(set(self.member_tags))) + 2) + ")"
        ws["U101"] = "=SUM(U3:U" + str(len(list(set(self.member_tags))) + 2) + ")"
        ws["V101"] = "=SUM(V3:V" + str(len(list(set(self.member_tags))) + 2) + ")"
        ws["W101"] = "=SUM(W3:W" + str(len(list(set(self.member_tags))) + 2) + ")"

        wb.save(filename)
        return None

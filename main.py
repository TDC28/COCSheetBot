import os
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()
warsL = []
playerTagL = []
starsL = []
desL = []
usernameL = []
wb = load_workbook("CWL2024.xlsx")
API_KEY = os.environ.get("API_KEY")


class COC:
    """Make calls to Clash of Clans API

    Parameters
    ----------
    token : 'str'
        token corresponding to SheetBot
    """

    def __init__(self, token: str | None) -> None:
        import requests

        self.requests = requests
        self.token = token
        self.api_endpoint = "https://api.clashofclans.com/v1"
        self.timeout = 45

    def get(self, uri, params=None):
        headers = {
            "Accept": "application/json",
            "Authorization": "Bearer " + self.token,
        }

        url = self.api_endpoint + uri

        try:
            response = self.requests.get(
                url, params=params, headers=headers, timeout=30
            )
            return response.json()

        except UnboundLocalError:
            return "Could not load response from Clash of Clans API"

    def leaguegroup(self, tag: str):
        """Returns the league group data associated to a given clan

        Parameters
        ----------
        tag : 'str'
            your clan tag"""
        return self.get(f"/clans/%23{tag}/currentwar/leaguegroup")

    def cwlwars(self, wartag: str, tag: str):
        """Gets clan war leages data from Clash of Clans API

        Parameters
        ----------
        wartag : 'str'
            tag corresponding to one war in the clan war league season

        tag : 'str'
            your clan tag
        """
        wars = self.get(f"/clanwarleagues/wars/%23{wartag}")

        if wars == {"state": "notInWar", "teamSize": 0}:
            pass

        elif wars["clan"]["tag"] != tag and wars["opponent"]["tag"] != tag:
            pass

        else:
            warsL.append(wars)


def getdata(tag):
    # Finds the clan matching clantag
    LeagueGroup = client.leaguegroup(tag[1:])

    # Append wars to warsL
    for groups in LeagueGroup["rounds"]:
        for tags in groups["warTags"]:
            client.cwlwars(tags[1:], tag)

    # Append player tag to playerTagL and name to usernameL
    for war in warsL:
        for player in range(0, warsL[0]["teamSize"]):
            if war["clan"]["tag"] == tag:
                playerTagL.append((war["clan"]["members"][player]["tag"]))
                usernameL.append((war["clan"]["members"][player]["name"]))

            else:
                playerTagL.append((war["opponent"]["members"][player]["tag"]))
                usernameL.append((war["opponent"]["members"][player]["name"]))

    # Append stars to starsL and destruction percentage to desL
    for war in warsL:
        for player in range(warsL[0]["teamSize"]):
            if war["clan"]["tag"] == tag:
                if "attacks" in war["clan"]["members"][player]:
                    starsL.append(
                        (war["clan"]["members"][player]["attacks"][0]["stars"])
                    )
                    desL.append(
                        (
                            war["clan"]["members"][player]["attacks"][0][
                                "destructionPercentage"
                            ]
                        )
                    )
                else:
                    starsL.append("Missed Hit")
                    desL.append(0)

            elif "attacks" in war["opponent"]["members"][player]:

                starsL.append(
                    (war["opponent"]["members"][player]["attacks"][0]["stars"])
                )
                desL.append(
                    (
                        war["opponent"]["members"][player]["attacks"][0][
                            "destructionPercentage"
                        ]
                    )
                )
            else:
                starsL.append(("Missed Hit"))
                desL.append(0)


def updatesheet():
    ws = wb[warsL[0]["startTime"][4:6]]

    for row in range(len(playerTagL)):
        ws["A" + str(row + 3)] = playerTagL[row]
        ws["B" + str(row + 3)] = starsL[row]
        ws["C" + str(row + 3)] = desL[row]
        ws["D" + str(row + 3)] = usernameL[row]

    ws["E3"] = warsL[0]["teamSize"]

    for row in range(0, len(list(set(playerTagL)))):
        ws["G" + str(row + 3)] = list(set(playerTagL))[row]
        ws["H" + str(row + 3)] = (
            "=_xlfn.XLOOKUP(G"
            + str(row + 3)
            + ",A2:A"
            + str(len(playerTagL) + 1)
            + ",D2:D"
            + str(len(playerTagL) + 1)
            + ")"
        )

    for row in range(len(set(playerTagL)) + 3, 101):
        ws.row_dimensions[row].hidden = True

    ws["S101"] = "=SUM(S3:S" + str(len(list(set(playerTagL))) + 2) + ")"
    ws["T101"] = "=SUM(T3:T" + str(len(list(set(playerTagL))) + 2) + ")"
    ws["U101"] = "=SUM(U3:U" + str(len(list(set(playerTagL))) + 2) + ")"
    ws["V101"] = "=SUM(V3:V" + str(len(list(set(playerTagL))) + 2) + ")"
    ws["W101"] = "=SUM(W3:W" + str(len(list(set(playerTagL))) + 2) + ")"

    wb.save("CWL2024.xlsx")


def otherdata(tag):
    # Finds the clan matching clantag
    LeagueGroup = client.leaguegroup(tag[1:])

    # Append wars to warsL
    for groups in LeagueGroup["rounds"]:
        for tags in groups["warTags"]:
            client.cwlwars(tags[1:], tag)

    # Append player tag to playerTagL and name to usernameL
    for war in warsL:
        for player in range(0, warsL[0]["teamSize"]):
            if war["clan"]["tag"] == tag:
                playerTagL.append((war["clan"]["members"][player]["tag"]))
                usernameL.append((war["clan"]["members"][player]["name"]))

            else:
                playerTagL.append((war["opponent"]["members"][player]["tag"]))
                usernameL.append((war["opponent"]["members"][player]["name"]))

    # Append stars to starsL and destruction percentage to desL
    for war in warsL:
        for player in range(warsL[0]["teamSize"]):
            if war["clan"]["tag"] == tag:
                if "attacks" in war["clan"]["members"][player]:
                    starsL.append(
                        (war["clan"]["members"][player]["attacks"][0]["stars"])
                    )
                    desL.append(
                        (
                            war["clan"]["members"][player]["attacks"][0][
                                "destructionPercentage"
                            ]
                        )
                    )
                else:
                    starsL.append("Missed Hit")
                    desL.append(0)

            elif "attacks" in war["opponent"]["members"][player]:

                starsL.append(
                    (war["opponent"]["members"][player]["attacks"][0]["stars"])
                )
                desL.append(
                    (
                        war["opponent"]["members"][player]["attacks"][0][
                            "destructionPercentage"
                        ]
                    )
                )
            else:
                starsL.append(("Missed Hit"))
                desL.append(0)

    ws = wb["Others" + warsL[0]["startTime"][4:6]]

    ws.append(playerTagL)
    ws.append(usernameL)
    ws.append(starsL)
    ws.append(desL)

    warsL.clear()
    playerTagL.clear()
    starsL.clear()
    desL.clear()
    usernameL.clear()
    wb.save("CWL2024.xlsx")


client = COC(API_KEY)
